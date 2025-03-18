import telebot
from openpyxl import load_workbook
from datetime import datetime
import os
import sys

# Телеграм-токен бота и ID администратора
TOKEN = "7925085246:AAFmHuGkoopznWShwkw-Eh745_ue6OdGTeY"
ADMIN_ID = 2133609169
bot = telebot.TeleBot(TOKEN)

# Определяем базовую директорию
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Названия файлов
BLACKLIST_FILE = os.path.join(BASE_DIR, "blacklist.xlsx")
PRODUCTS_FILE = os.path.join(BASE_DIR, "products.xlsx")
ORDERS_FILE = os.path.join(BASE_DIR, "orders.xlsx")

# Уведомление админа
def notify_admin(order_details):
    try:
        bot.send_message(ADMIN_ID, f"🔔 Новый заказ!\n{order_details}")
    except Exception as e:
        print(f"Ошибка отправки уведомления админу: {e}")

# /start и /help — список команд
@bot.message_handler(commands=['start', 'help'])
def send_help(message):
    help_text = (
        "/help - список доступных команд\n"
        "/Magaz - показать ассортимент магазина\n"
        "/Order - оформить заказ\n"
        "/blacklist - показать черный список персонажей"
    )
    if message.from_user.id == ADMIN_ID:
        help_text += (
            "\n\n*Команды администратора:*\n"
            "/Orders - посмотреть все заказы\n"
            "/AddToBlacklist - добавить в черный список"
        )
    bot.reply_to(message, help_text, parse_mode="Markdown")

# /Magaz — показать ассортимент
@bot.message_handler(commands=['Magaz'])
def send_products(message):
    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    categories = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4 or any(cell is None for cell in row):
            continue
        name, category, price, qty = row
        categories.setdefault(category, []).append((name, price, qty))

    if not categories:
        bot.send_message(message.chat.id, "❌ Магазин пуст! Проверьте данные в файле.")
        return

    lines = []
    for cat in sorted(categories.keys()):
        lines.append(f"*{cat}*:")
        for name, price, qty in categories[cat]:
            lines.append(f" - {name}: {price} (в наличии: {qty})")
        lines.append("")

    text = "\n".join(lines)  # Исправлено здесь
    bot.send_message(message.chat.id, text, parse_mode="Markdown")

# /blacklist — показать черный список
@bot.message_handler(commands=['ЧС', 'blacklist', 'banlist'])
def send_blacklist(message):
    try:
        wb = load_workbook(BLACKLIST_FILE)
        ws = wb.active
        lines = ["*Черный список персонажей:*"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"DEBUG: {row}")
            if len(row) < 2 or any(cell is None for cell in row[:2]):
                continue
            name, reason = row[:2]
            lines.append(f" - {name} ({reason})")

        if len(lines) == 1:
            text = "🔹 Черный список пуст."
        else:
            text = "\n".join(lines)

        bot.send_message(message.chat.id, text, parse_mode="Markdown")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка чтения черного списка: {e}")
        print(f"Ошибка: {e}")

# /Order — оформить заказ
@bot.message_handler(commands=['Order'])
def order_start(message):
    msg = bot.reply_to(message, "Введите название товара для заказа:")
    bot.register_next_step_handler(msg, process_order_product, message.from_user.id)

def process_order_product(message, user_id):
    product_name = message.text.strip()

    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    product_name_found = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]) if row[0] else ""
        if name.lower() == product_name.lower():
            product_name_found = name
            break

    if product_name_found is None:
        bot.reply_to(message, "❌ Товар не найден. Заказ отменен. Используйте /Order для новой попытки.")
        return

    msg = bot.reply_to(message, f'Введите количество для товара "{product_name_found}":')
    bot.register_next_step_handler(msg, process_order_quantity, product_name_found, user_id)

def process_order_quantity(message, product_name, user_id):
    qty_text = message.text.strip()

    if not qty_text.isdigit():
        msg = bot.reply_to(message, "Количество должно быть числом. Введите количество:")
        bot.register_next_step_handler(msg, process_order_quantity, product_name, user_id)
        return

    quantity = int(qty_text)
    if quantity <= 0:
        msg = bot.reply_to(message, "Количество должно быть больше нуля. Введите количество:")
        bot.register_next_step_handler(msg, process_order_quantity, product_name, user_id)
        return

    wb = load_workbook(PRODUCTS_FILE)
    ws = wb.active
    target_row = None
    for row in ws.iter_rows(min_row=2):
        name_cell = row[0]
        qty_cell = row[3]
        if name_cell.value and str(name_cell.value).lower() == product_name.lower():
            target_row = row
            break

    if target_row is None:
        bot.reply_to(message, "Ошибка: товар не найден.")
        return

    name_cell, cat_cell, price_cell, qty_cell = target_row
    available = qty_cell.value or 0

    if quantity > available:
        bot.reply_to(message, f'Недостаточно товара "{name_cell.value}" в наличии (доступно {available}).')
        return

    qty_cell.value = available - quantity
    wb.save(PRODUCTS_FILE)

    username = message.from_user.username or f"{message.from_user.first_name} {message.from_user.last_name}".strip()
    if not username:
        username = f"ID: {message.from_user.id}"

    wb_orders = load_workbook(ORDERS_FILE)
    ws_orders = wb_orders.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_orders.append([now, username, name_cell.value, quantity])
    wb_orders.save(ORDERS_FILE)

    order_details = (f"👤 Покупатель: {username}\n"
                    f"📦 Товар: {name_cell.value}\n"
                    f"📊 Количество: {quantity}\n"
                    f"🕒 Время: {now}")

    bot.send_message(user_id, f"✅ Ваш заказ оформлен: {name_cell.value} x {quantity}. Спасибо!")
    notify_admin(order_details)

    bot.reply_to(message, "Заказ принят! Ожидайте подтверждения.")

# /Orders — показать все заказы (только для админа)
@bot.message_handler(commands=['Orders'])
def show_orders(message):
    if message.from_user.id != ADMIN_ID:
        bot.reply_to(message, "❌ Доступ запрещен. Эта команда только для администратора.")
        return

    try:
        wb = load_workbook(ORDERS_FILE)
        ws = wb.active
        lines = ["*Список заказов:*"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4 or any(cell is None for cell in row):
                continue
            date, username, product, qty = row
            date = str(date).replace('_', r'\_')
            username = str(username).replace('_', r'\_')
            product = str(product).replace('_', r'\_')
            qty = str(qty).replace('_', r'\_')
            line = f"🕒 {date} | 👤 {username} | 📦 {product} x{qty}"
            if len("\n".join(lines + [line])) < 4000:
                lines.append(line)
            else:
                break

        if len(lines) == 1:
            text = "🔹 Список заказов пуст."
        else:
            text = "\n".join(lines)

        try:
            bot.send_message(message.chat.id, text, parse_mode="Markdown")
        except telebot.apihelper.ApiTelegramException as e:
            bot.send_message(message.chat.id, text, parse_mode=None)
            print(f"Ошибка Markdown: {e}")
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка при чтении заказов: {e}")

# /AddToBlacklist — добавить в черный список (только для админа)
@bot.message_handler(commands=['AddToBlacklist', 'ДобавитьВЧС'])
def add_to_blacklist_start(message):
    if message.from_user.id != ADMIN_ID:
        bot.reply_to(message, "❌ Доступ запрещен. Эта команда только для администратора.")
        return
    
    msg = bot.reply_to(message, "Введите имя для добавления в черный список:")
    bot.register_next_step_handler(msg, process_blacklist_name)

def process_blacklist_name(message):
    name = message.text.strip()
    msg = bot.reply_to(message, f"Введите причину для '{name}' в черном списке:")
    bot.register_next_step_handler(msg, process_blacklist_reason, name)

def process_blacklist_reason(message, name):
    reason = message.text.strip()
    
    try:
        wb = load_workbook(BLACKLIST_FILE)
        ws = wb.active
        ws.append([name, reason])
        wb.save(BLACKLIST_FILE)
        bot.reply_to(message, f"✅ '{name}' добавлен в черный список с причиной: {reason}")
        notify_admin(f"🔔 Добавлен в ЧС:\n👤 {name}\n📝 Причина: {reason}")
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка при добавлении в черный список: {e}")

# Запуск бота
bot.infinity_polling()